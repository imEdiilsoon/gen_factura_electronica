import os
import random
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from openpyxl import load_workbook, Workbook

opcion = -1

while opcion != 2:
  print("\nElige una opción: \n\n | 1 - Crear una nueva Factura. \n | 2 - Salir del programa. \n")
  opcion = int(input("→ "))
  
  if opcion == 2:
    os.system("cls")
    exit

  productos = []
  precios = []
  cantidades = []

  if opcion == 1:
    os.system("cls")
    opcion_menu_factura = -1
    id_factura = f"Fact-{random.randint(1000, 9000)}"

    print("| 👤 Información básica del cliente.")
    print("\nIngrese el nombre del cliente:")
    cliente = input("→ ")
    print("\nIngrese la cedula del cliente:")
    cedula = int(input("→ "))
    print("\nIngrese el correo del cliente:")
    correo = input("→ ")
    print("\nIngrese el número de teléfono del cliente:")
    numero_telf = input("→ ")

    while opcion_menu_factura != 2:

      os.system("cls")
      print("| 🛒 Información del producto.")
      print("\n Ingrese el nombre del producto:")
      productos.append(input("→ "))
      print("\n Ingrese el valor unitario del producto:")
      precios.append(int(input("$ ")))
      print("\n Ingrese la cantidad del producto:")
      cantidades.append(int(input("→ ")))

      os.system("cls")
      print("\nElige una opción: \n\n | 1 - Registrar otro producto. \n | 2 - Imprimir factura. \n")
      opcion_menu_factura = int(input("→ "))

      if(opcion_menu_factura == 2):
        os.system("cls")
        print("__________________________________________________________")
        print(f"\n|  🧾 Factura Generada N°: {id_factura}\n")
        
        for i in range(len(productos)):
          print(f"✅ Producto: {productos[i]} | Precio Unit: {precios[i]} | Cantidad: {cantidades[i]}")

        subtotal = 0
        for i in range(len(precios)):
          subtotal = subtotal + (precios[i] * cantidades[i])

        iva = int(subtotal * .19)
        total = int(iva + subtotal)
        fecha = datetime.now().strftime('%d/%m/%Y - %H:%M:%S')

        descuento = False
        if(len(productos) > 6):
          descuento = True

        print(f"\nSubtotal: $ {subtotal}")
        print(f"IVA (19%): $ {iva}")
        if descuento:
          print(f"Dto aplicado: 10%")
          print(f"Total: $ {int(total - (total * 0.1))}")
        else:
          print(f"Total: $ {total}")

        print("\n|  👤 Información del cliente:\n")
        print(f"Nombre: {cliente}.")
        print(f"Cedula: {cedula}.")
        print(f"Correo: {correo}.")
        print(f"Tlf: {numero_telf}.")
        print("\nGracias por tu compra! ❤️\n")
        print(f"|  📅 Fecha: {fecha}")
        print("__________________________________________________________")

        opcion_pdf_gen = -1
        print("\n¿Desea guardar esta factura en un PDF?\n\n | 1 - Si.\n | 2 - No.\n")
        opcion_pdf_gen = int(input("→ "))

        if opcion_pdf_gen == 1:
          def Generar_factura(nombre_archivo):
            ruta_guardar_pdf= "pdf_generados"

            if os.path.exists(ruta_guardar_pdf):
              ruta_completa = os.path.join(ruta_guardar_pdf, nombre_archivo)
            else:
              os.mkdir("pdf_generados")
              ruta_guardar_pdf = "pdf_generados"
              ruta_completa = os.path.join(ruta_guardar_pdf, nombre_archivo)
            
            ancho = 120 * mm
            alto = 300 * mm
            c = canvas.Canvas(ruta_completa, pagesize=(ancho, alto))
            
            text = c.beginText()
            text.setTextOrigin(50, 750)
            text.setFont("Courier", 12)

            text.textLine(f"Factura n°: {id_factura}")
            text.textLine(" ")
            text.textLine(" ")
            text.textLine("Información de la compra:")
            text.textLine(" ")

            for i in range(len(productos)):
              valor_neto = precios[i] * cantidades[i]
              text.textLine(f"{productos[i]} x ({cantidades[i]})  - $ {valor_neto}")

            text.textLine(" ")
            text.textLine(f"Subtotal: $ {subtotal}")
            text.textLine(f"IVA (19%): $ {iva}")
            if descuento:
              text.textLine(f"Dto aplicado: 10%")
              text.textLine(f"Total: $ {int(total - (total * .1))}")
            else:
              text.textLine(f"Total: $ {total}")
            text.textLine(" ")
            text.textLine(" ")

            text.textLine("Información del cliente:")
            text.textLine(" ")
            text.textLine(f"Nombre: {cliente}")
            text.textLine(f"Cedula: {cedula}")
            text.textLine(f"Correo: {correo}")
            text.textLine(f"Tlf: {numero_telf}")
            text.textLine(" ")
            text.textLine("¡Gracias por tu compra!")
            text.textLine(" ")
            text.textLine(f"Fecha: {fecha}")
            
            c.drawText(text)
            c.showPage()
            c.save()
            print(f"\nLa Factura n°: {id_factura}. Se guardó correctamente.")

          Generar_factura(f"{id_factura}.pdf")

        elif opcion_pdf_gen == 2:
          os.system("cls")

        archivo_excel = "facturas.xlsx"

        if os.path.exists(archivo_excel):
          wb = load_workbook(archivo_excel)
          ws = wb.active
        else:
          wb = Workbook()
          ws = wb.active
          ws.append(["id_factura", "nombre_producto", "valor_unit_producto", "cantidad_producto", "valor_neto_producto", "iva_producto", "descuento_producto", "precio_total_producto", "nombre_cliente", "cedula_cliente", "correo_cliente", "telefono_cliente", "fecha_compra"])

        for i in range(len(productos)):
          valor_neto_producto = precios[i] * cantidades[i]
          iva_producto = valor_neto_producto * 0.19
          precio_total_producto = valor_neto_producto + iva_producto
          descuento_producto = 0
          if descuento:
            descuento_producto = precio_total_producto * .1
            precio_total_producto = precio_total_producto - descuento_producto
        
          ws.append([id_factura, productos[i], precios[i], cantidades[i], valor_neto_producto, iva_producto, descuento_producto, precio_total_producto, cliente, cedula, correo, numero_telf, fecha])

        wb.save(archivo_excel)
